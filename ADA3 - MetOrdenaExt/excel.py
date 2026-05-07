"""
Ordenamiento Externo con Archivos Excel
=========================================
Simula cintas magnéticas guardando datos en hojas de un libro Excel.

Archivos generados en la carpeta  cintas/  junto al script:
  - cintas.xlsx        → libro de trabajo con una hoja por cinta
  - resultado.xlsx     → 3 hojas de presentación del resultado:
        Hoja 1 "Vertical"    → datos en columna  (A1:A n)
        Hoja 2 "Horizontal"  → datos en fila     (A1:n1)
        Hoja 3 "Revuelto"    → datos desordenados en cuadrícula

Compatible: Python 3.8+  |  tkinter (stdlib) + openpyxl
"""

import tkinter as tk
from tkinter import filedialog, scrolledtext
import random
import threading
import os
import time
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

# ══════ PALETA UI ══════
BG      = "#0a0d12"
PANEL   = "#111720"
PANEL2  = "#161d2a"
BORDER  = "#1e2d40"
C_CMP   = "#00d4ff"
C_SWP   = "#ff6b6b"
C_AUX   = "#a78bfa"
C_DONE  = "#22c55e"
C_BAR   = "#1e2d40"
C_READ  = "#f59e0b"
TXT_HI  = "#e2e8f0"
TXT_LO  = "#64748b"
TXT_MID = "#94a3b8"
BTN_BG  = "#1e2d40"
C_ERR   = "#ff6b6b"
C_WARN  = "#f59e0b"

CINTA_COLORES = {"A":"#00d4ff","B":"#a78bfa","C":"#22c55e","D":"#f59e0b"}

# ══════ ESTILOS EXCEL ══════
# Colores de fondo para cada cinta en el Excel
XLSX_FILLS = {
    "cinta_A": "D6F0FB",
    "cinta_B": "EDE9FF",
    "cinta_C": "D4F4E2",
    "cinta_D": "FEF3C7",
    "datos":   "F1F5F9",
    "entrada": "F1F5F9",
    "temp1":   "FFF1F0",
    "temp2":   "FFF8EC",
}
HEADER_FILL = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT   = Font(name="Arial", size=9)
THIN        = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

# ══════ UTILIDADES DE ARCHIVO ══════

def carpeta_cintas():
    base = os.path.dirname(os.path.abspath(__file__))
    ruta = os.path.join(base, "cintas")
    os.makedirs(ruta, exist_ok=True)
    return ruta

def ruta(nombre):
    return os.path.join(carpeta_cintas(), nombre)

# ── Libro compartido para cintas intermedias ────────────────

def _get_libro_cintas():
    """Abre o crea cintas.xlsx."""
    path = ruta("cintas.xlsx")
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception:
            pass
    return openpyxl.Workbook()  # hoja Sheet por defecto

def escribir_cinta(nombre, numeros):
    """
    Escribe la lista en una hoja del libro cintas.xlsx.
    nombre puede ser  'cinta_A', 'datos', 'resultado', etc.
    """
    path = ruta("cintas.xlsx")
    wb   = _get_libro_cintas()

    # Crear o limpiar hoja
    hoja_nombre = nombre.replace(".txt", "").replace(".xlsx", "")
    if hoja_nombre in wb.sheetnames:
        del wb[hoja_nombre]
    ws = wb.create_sheet(hoja_nombre)

    # Cabecera
    fill_color = XLSX_FILLS.get(hoja_nombre, "F8FAFC")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14

    ws["A1"] = "#"
    ws["B1"] = "Valor"
    for cell in (ws["A1"], ws["B1"]):
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = CELL_BORDER

    # Datos
    row_fill = _fill(fill_color)
    for i, val in enumerate(numeros, start=1):
        ws[f"A{i+1}"] = i
        ws[f"B{i+1}"] = val
        for col in ("A", "B"):
            c = ws[f"{col}{i+1}"]
            c.fill      = row_fill
            c.font      = CELL_FONT
            c.border    = CELL_BORDER
            c.alignment = Alignment(horizontal="center")

    # Eliminar hoja por defecto si quedan otras
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    wb.save(path)

def leer_cinta(nombre):
    """Lee la columna B de la hoja correspondiente en cintas.xlsx."""
    path = ruta("cintas.xlsx")
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hoja_nombre = nombre.replace(".txt", "").replace(".xlsx", "")
    if hoja_nombre not in wb.sheetnames:
        return []
    ws  = wb[hoja_nombre]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            try:
                out.append(int(row[1]))
            except (ValueError, TypeError):
                pass
    return out

def tamanio_cinta(nombre):
    path = ruta("cintas.xlsx")
    return os.path.getsize(path) if os.path.exists(path) else 0

# ── Resultado final en resultado.xlsx (3 hojas) ─────────────

def escribir_resultado_excel(numeros):
    """
    Crea resultado.xlsx con 3 hojas:
      1. Vertical   → una columna
      2. Horizontal → una fila
      3. Revuelto   → cuadrícula desordenada
    """
    path = ruta("resultado.xlsx")
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)          # quitar hoja por defecto

    n = len(numeros)

    # ── HOJA 1: Vertical ────────────────────────────────────
    ws1 = wb.create_sheet("Vertical")
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16

    ws1["A1"] = "#"
    ws1["B1"] = "Valor ordenado"
    for c in (ws1["A1"], ws1["B1"]):
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border    = CELL_BORDER

    # Degradado verde para indicar orden
    greens = ["D4F4E2", "C6EFD3", "B7E9C4", "A8E3B5", "99DDA6"]
    for i, val in enumerate(numeros, start=1):
        fill_hex = greens[min(int(i / max(n,1) * len(greens)), len(greens)-1)]
        ws1[f"A{i+1}"] = i
        ws1[f"B{i+1}"] = val
        for col in ("A", "B"):
            c = ws1[f"{col}{i+1}"]
            c.fill      = _fill(fill_hex)
            c.font      = CELL_FONT
            c.border    = CELL_BORDER
            c.alignment = Alignment(horizontal="center")

    ws1.freeze_panes = "A2"

    # ── HOJA 2: Horizontal ──────────────────────────────────
    ws2 = wb.create_sheet("Horizontal")

    # Fila 1: encabezados (1, 2, 3, …)
    # Fila 2: valores
    for col_idx, val in enumerate(numeros, start=1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 8

        h_cell = ws2.cell(row=1, column=col_idx, value=col_idx)
        h_cell.fill      = HEADER_FILL
        h_cell.font      = HEADER_FONT
        h_cell.alignment = Alignment(horizontal="center")
        h_cell.border    = CELL_BORDER

        v_cell = ws2.cell(row=2, column=col_idx, value=val)
        blues  = ["D6EAF8","C0D8F0","AAC6E8","94B4E0","7EA2D8"]
        fill_hex = blues[min(int(col_idx / max(n,1) * len(blues)), len(blues)-1)]
        v_cell.fill      = _fill(fill_hex)
        v_cell.font      = CELL_FONT
        v_cell.border    = CELL_BORDER
        v_cell.alignment = Alignment(horizontal="center")

    ws2.row_dimensions[1].height = 18
    ws2.row_dimensions[2].height = 20

    # ── HOJA 3: Revuelto ────────────────────────────────────
    ws3 = wb.create_sheet("Revuelto")

    # Calcular cuántas columnas usar (cuadrícula más cuadrada posible)
    cols = max(1, min(20, int(n**0.5) + 1))
    rows = (n + cols - 1) // cols

    revuelto = numeros[:]
    random.shuffle(revuelto)          # desordenar

    # Título
    ws3.merge_cells("A1:" + get_column_letter(cols) + "1")
    t = ws3["A1"]
    t.value     = "Datos Revueltos (desordenados)"
    t.fill      = _fill("7C3AED")
    t.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    t.alignment = Alignment(horizontal="center")
    t.border    = CELL_BORDER

    # Paleta aleatoria de colores pastel
    paleta = ["FFF3CD","D4EDDA","D1ECF1","F8D7DA","E2D9F3",
              "FDE2E4","C3E6CB","BEE5EB","FFEEBA","D6E4FF"]

    idx = 0
    for r in range(rows):
        for c in range(cols):
            if idx >= len(revuelto):
                break
            col_letter = get_column_letter(c + 1)
            ws3.column_dimensions[col_letter].width = 9
            cell = ws3.cell(row=r + 2, column=c + 1, value=revuelto[idx])
            cell.fill      = _fill(paleta[(idx) % len(paleta)])
            cell.font      = CELL_FONT
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(horizontal="center")
            idx += 1
        ws3.row_dimensions[r + 2].height = 16

    wb.save(path)
    return path


# ══════ ALGORITMOS ══════
# (misma lógica, solo cambiada la E/S de txt → Excel)

class AlgoritmoBase:
    def __init__(self, log_cb, barra_cb, cinta_cb, fin_cb, pausa_fn):
        self.log   = log_cb
        self.barra = barra_cb
        self.cinta = cinta_cb
        self.fin   = fin_cb
        self.pausa = pausa_fn
        self._stop = False

    def detener(self):
        self._stop = True

    def _chk(self):
        if self._stop:
            raise InterruptedError
        self.pausa()

    def _mostrar_cintas(self, *nombres):
        d = {}
        for n in nombres:
            d[n] = leer_cinta(n)
        self.cinta(d)


class Intercalacion(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        self.log("═"*50, TXT_LO)
        self.log("INTERCALACIÓN — guardando en Excel", C_CMP)
        self.log("═"*50, TXT_LO)

        datos = random.choices(range(1, n_elem*10+1), k=n_elem)
        escribir_cinta("datos", datos)
        self.log(f"✦ Hoja 'datos' generada ({n_elem} elementos)", C_READ)
        self._mostrar_cintas("datos")
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()

        data, size, n, paso = datos[:], 1, len(datos), 0

        while size < n:
            self._chk()
            paso += 1
            self.log(f"\n── Pasada {paso} (bloque={size}) ──", TXT_MID)

            cA, cB, i, toggle = [], [], 0, True
            while i < n:
                bloque = data[i:i+size]
                (cA if toggle else cB).extend(bloque)
                toggle = not toggle
                i += size

            escribir_cinta("cinta_A", cA)
            escribir_cinta("cinta_B", cB)
            self.log(f"  → Hoja cinta_A ({len(cA)}) | cinta_B ({len(cB)})", C_AUX)
            self._mostrar_cintas("cinta_A", "cinta_B")
            time.sleep(delay)
            self._chk()

            resultado, left = [], 0
            while left < n:
                mid   = min(left+size,   n)
                right = min(left+size*2, n)
                i2, j = left, mid
                while i2 < mid and j < right:
                    self._chk()
                    self.barra(data, {i2: C_CMP, j: C_CMP})
                    time.sleep(delay*0.3)
                    if data[i2] <= data[j]:
                        resultado.append(data[i2]); i2 += 1
                    else:
                        resultado.append(data[j]);  j  += 1
                while i2 < mid:
                    resultado.append(data[i2]); i2 += 1
                while j < right:
                    resultado.append(data[j]);  j  += 1
                left += size*2

            data = resultado
            escribir_cinta("datos", data)
            self.log(f"  ✔ fusión → hoja 'datos' actualizada", C_DONE)
            self.barra(data, {})
            self._mostrar_cintas("datos")
            time.sleep(delay)
            size *= 2

        path = escribir_resultado_excel(data)
        self.log(f"\n✔ resultado.xlsx escrito", C_DONE)
        self._mostrar_cintas("datos")
        self.fin(data, path)


class MezclaDirecta(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        self.log("═"*50, TXT_LO)
        self.log("MEZCLA DIRECTA — guardando en Excel", C_CMP)
        self.log("═"*50, TXT_LO)

        datos = random.choices(range(1, n_elem*10+1), k=n_elem)
        escribir_cinta("entrada", datos)
        self.log(f"✦ Hoja 'entrada' generada ({n_elem} elementos)", C_READ)
        self._mostrar_cintas("entrada")
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()

        data, steps = datos[:], []

        def merge(lo, mid, hi):
            left  = data[lo:mid+1]
            right = data[mid+1:hi+1]
            i=j=0; k=lo
            while i<len(left) and j<len(right):
                if left[i]<=right[j]: data[k]=left[i]; i+=1
                else:                 data[k]=right[j]; j+=1
                steps.append(("merge", data[:], lo, mid, hi, k)); k+=1
            while i<len(left):
                data[k]=left[i]; i+=1; k+=1
                steps.append(("merge", data[:], lo, mid, hi, k-1))
            while j<len(right):
                data[k]=right[j]; j+=1; k+=1
                steps.append(("merge", data[:], lo, mid, hi, k-1))

        def sort(lo, hi):
            if lo>=hi: return
            mid=(lo+hi)//2
            steps.append(("div", data[:], lo, mid, hi, 0))
            sort(lo, mid); sort(mid+1, hi); merge(lo, mid, hi)
            steps.append(("merged", data[:], lo, mid, hi, 0))

        sort(0, len(data)-1)

        for ev, snap, lo, mid, hi, extra in steps:
            self._chk()
            if ev=="div":
                self.barra(snap, {k:C_AUX for k in range(lo,hi+1)})
                self.log(f"  División [{lo}..{mid}] [{mid+1}..{hi}]", C_AUX)
                escribir_cinta("temp1", snap[lo:mid+1])
                escribir_cinta("temp2", snap[mid+1:hi+1])
                self._mostrar_cintas("temp1","temp2")
                time.sleep(delay*0.5)
            elif ev=="merge":
                self.barra(snap, {extra:C_SWP, lo:C_CMP, hi:C_CMP})
                escribir_cinta("entrada", snap)
                time.sleep(delay*0.2)
            elif ev=="merged":
                self.barra(snap, {k:C_DONE for k in range(lo,hi+1)})
                self.log(f"  ✔ Fusión [{lo}..{hi}]", C_DONE)
                self._mostrar_cintas("entrada")
                time.sleep(delay*0.4)

        path = escribir_resultado_excel(data)
        self.log(f"\n✔ resultado.xlsx escrito", C_DONE)
        self.fin(data, path)


class MezclaEquilibrada(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        self.log("═"*50, TXT_LO)
        self.log("MEZCLA EQUILIBRADA — 4 hojas Excel", C_CMP)
        self.log("═"*50, TXT_LO)

        datos = random.choices(range(1, n_elem*10+1), k=n_elem)
        escribir_cinta("datos", datos)
        self.log(f"✦ Hoja 'datos' generada ({n_elem} elementos)", C_READ)

        cA = datos[0::2]; cB = datos[1::2]
        escribir_cinta("cinta_A", cA)
        escribir_cinta("cinta_B", cB)
        self.log(f"✦ cinta_A: {len(cA)} | cinta_B: {len(cB)}", C_AUX)
        self._mostrar_cintas("cinta_A","cinta_B")
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()

        data = datos[:]
        n, size, pasada = len(data), 1, 0
        pares = [("cinta_A","cinta_B","cinta_C","cinta_D"),
                 ("cinta_C","cinta_D","cinta_A","cinta_B")]

        while size < n:
            self._chk()
            ent1,ent2,sal1,sal2 = pares[pasada%2]
            pasada += 1
            self.log(f"\n── Pasada {pasada} (bloque={size})", TXT_MID)
            self.log(f"   Lee: {ent1} + {ent2}", C_READ)
            self.log(f"   Escribe: {sal1} + {sal2}", C_SWP)

            inp1, inp2 = leer_cinta(ent1), leer_cinta(ent2)

            # Reconstruir data_ent intercalando bloques
            data_ent, i1, i2, toggle = [], 0, 0, True
            while i1<len(inp1) or i2<len(inp2):
                if toggle and i1<len(inp1):
                    data_ent.extend(inp1[i1:i1+size]); i1+=size
                elif not toggle and i2<len(inp2):
                    data_ent.extend(inp2[i2:i2+size]); i2+=size
                else:
                    if i1<len(inp1):   data_ent.extend(inp1[i1:i1+size]); i1+=size
                    elif i2<len(inp2): data_ent.extend(inp2[i2:i2+size]); i2+=size
                toggle = not toggle

            out1, out2, left, t2 = [], [], 0, True
            while left < len(data_ent):
                mid2  = min(left+size,       len(data_ent))
                right = min(left+size*2,     len(data_ent))
                ia, jb, merged = left, mid2, []
                while ia<mid2 and jb<right:
                    self._chk()
                    self.barra(data_ent, {ia:C_CMP, jb:C_CMP})
                    time.sleep(delay*0.15)
                    if data_ent[ia]<=data_ent[jb]: merged.append(data_ent[ia]); ia+=1
                    else:                           merged.append(data_ent[jb]); jb+=1
                while ia<mid2:  merged.append(data_ent[ia]); ia+=1
                while jb<right: merged.append(data_ent[jb]); jb+=1
                (out1 if t2 else out2).extend(merged)
                t2 = not t2; left += size*2

            # Reordenar intercalando
            resultado, i1, i2, tog = [], 0, 0, True
            while i1<len(out1) or i2<len(out2):
                bs=size*2
                if tog and i1<len(out1):   resultado.extend(out1[i1:i1+bs]); i1+=bs
                elif not tog and i2<len(out2): resultado.extend(out2[i2:i2+bs]); i2+=bs
                else:
                    if i1<len(out1):   resultado.extend(out1[i1:i1+bs]); i1+=bs
                    elif i2<len(out2): resultado.extend(out2[i2:i2+bs]); i2+=bs
                tog = not tog
            data = resultado or out1+out2

            escribir_cinta(sal1, out1)
            escribir_cinta(sal2, out2)
            self.log(f"   {sal1}: {len(out1)} | {sal2}: {len(out2)}", C_DONE)
            self._mostrar_cintas(sal1, sal2)
            self.barra(data, {})
            time.sleep(delay)
            size *= 2

        final = sorted(leer_cinta(pares[(pasada-1)%2][2]) +
                       leer_cinta(pares[(pasada-1)%2][3]))
        path = escribir_resultado_excel(final)
        self.log(f"\n✔ resultado.xlsx escrito", C_DONE)
        self.barra(final, {i:C_DONE for i in range(len(final))})
        self.fin(final, path)


# ══════ WIDGETS UI ══════

class TarjetaCinta(tk.Frame):
    def __init__(self, master, nombre, color, **kw):
        super().__init__(master, bg=PANEL2, highlightthickness=1,
                         highlightbackground=color, **kw)
        self.color  = color
        self.nombre = nombre
        tk.Label(self, text=f"📊 {nombre}", bg=PANEL2,
                 fg=color, font=("Courier New",9,"bold")).pack(anchor="w",padx=6,pady=(4,0))
        self.lbl_info = tk.Label(self, text="vacía", bg=PANEL2, fg=TXT_LO,
                                 font=("Courier New",8))
        self.lbl_info.pack(anchor="w", padx=6)
        self.txt = tk.Text(self, height=3, bg="#080c10", fg=color,
                           font=("Courier New",8), state="disabled",
                           relief="flat", bd=0, wrap="word")
        self.txt.pack(fill="x", padx=4, pady=(2,4))

    def actualizar(self, datos):
        n = len(datos)
        kb = tamanio_cinta(self.nombre) / 1024
        self.lbl_info.config(text=f"{n} filas · cintas.xlsx: {kb:.1f} KB")
        self.txt.config(state="normal"); self.txt.delete("1.0","end")
        preview = datos[:40]
        txt = " ".join(str(x) for x in preview)
        if len(datos)>40: txt += f"  … (+{len(datos)-40})"
        self.txt.insert("end", txt); self.txt.config(state="disabled")

    def limpiar(self):
        self.lbl_info.config(text="vacía")
        self.txt.config(state="normal"); self.txt.delete("1.0","end")
        self.txt.config(state="disabled")


class BarrasCanvas(tk.Canvas):
    def __init__(self, master, **kw):
        super().__init__(master, bg=BG, height=160,
                         highlightthickness=1, highlightbackground=BORDER, **kw)

    def dibujar(self, arr, highlights=None):
        hi = highlights or {}
        self.delete("all")
        W = self.winfo_width()  or 800
        H = self.winfo_height() or 160
        n = len(arr)
        if not n: return
        GAP   = 2
        bar_w = max(2,(W-GAP*(n+1))//n)
        maxv  = max(arr) or 1
        PAD_B = 18
        for i,v in enumerate(arr):
            x1=GAP+i*(bar_w+GAP); x2=x1+bar_w
            bh=max(3,int((v/maxv)*(H-PAD_B-8)))
            y2=H-PAD_B; y1=y2-bh
            col=hi.get(i,C_BAR)
            self.create_rectangle(x1,y1,x2,y2,fill=col,outline="")
            if bar_w>=14:
                self.create_text(x1+bar_w//2,y1-6,text=str(v),
                                 fill=TXT_LO,font=("Courier New",7))
        ley=[("Comparando",C_CMP),("Escribiendo",C_SWP),
             ("Aux",C_AUX),("Ordenado",C_DONE),("Leyendo",C_READ)]
        lx=6; ly=H-13
        for etq,col in ley:
            self.create_rectangle(lx,ly,lx+10,ly+9,fill=col,outline="")
            self.create_text(lx+14,ly+4,text=etq,anchor="w",
                             fill=TXT_LO,font=("Courier New",7))
            lx+=88


# ══════ VENTANA PRINCIPAL ══════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Ordenamiento Externo → Excel — Visualizador")
        self.configure(bg=BG)
        self.geometry("1000x900")
        self.minsize(800,700)
        self._algo     = None
        self._thread   = None
        self._pausado  = threading.Event()
        self._pausado.set()
        self._delay    = 0.3
        self._datos_externos = None
        self._build()

    def _build(self):
        tk.Frame(self, bg=C_CMP, height=3).pack(fill="x")

        hdr=tk.Frame(self,bg=BG); hdr.pack(fill="x",padx=16,pady=10)
        tk.Label(hdr,text="Ordenamiento Externo → Excel",
                 font=("Segoe UI",15,"bold"),bg=BG,fg=TXT_HI).pack(side="left")
        tk.Label(hdr,text="cintas.xlsx  +  resultado.xlsx",
                 font=("Courier New",9),bg=BG,fg=TXT_LO).pack(side="right",anchor="s")
        tk.Frame(self,bg=BORDER,height=1).pack(fill="x")

        # Config
        cfg=tk.Frame(self,bg=PANEL,pady=8); cfg.pack(fill="x",padx=12,pady=(8,4))
        tk.Label(cfg,text="Elementos:",bg=PANEL,fg=TXT_HI,
                 font=("Segoe UI",10,"bold")).pack(side="left",padx=(10,4))
        self.n_var=tk.StringVar(value="20")
        tk.Entry(cfg,textvariable=self.n_var,width=7,
                 bg="#0a0d12",fg=C_CMP,font=("Courier New",12,"bold"),
                 relief="flat",bd=4,justify="center",
                 insertbackground=C_CMP,highlightthickness=2,
                 highlightbackground=BORDER,highlightcolor=C_CMP
                 ).pack(side="left",padx=4)
        for v in (10,50,100,500,1000):
            tk.Button(cfg,text=str(v),bg=BTN_BG,fg=TXT_LO,
                      activebackground=C_CMP,activeforeground=BG,
                      relief="flat",font=("Courier New",8,"bold"),
                      padx=6,pady=2,cursor="hand2",bd=0,
                      command=lambda x=v:self.n_var.set(str(x))
                      ).pack(side="left",padx=2)

        tk.Frame(cfg,bg=BORDER,width=1,height=28).pack(side="left",padx=10)
        tk.Label(cfg,text="Velocidad:",bg=PANEL,fg=TXT_HI,
                 font=("Segoe UI",10,"bold")).pack(side="left",padx=(4,4))
        self.vel_var=tk.IntVar(value=50)
        tk.Scale(cfg,from_=1,to=100,variable=self.vel_var,orient="horizontal",
                 length=110,bg=PANEL,fg=TXT_LO,troughcolor=BORDER,
                 highlightthickness=0,showvalue=False,
                 command=self._vel_changed).pack(side="left",padx=4)

        tk.Frame(cfg,bg=BORDER,width=1,height=28).pack(side="left",padx=10)
        tk.Button(cfg,text="📂 Cargar .txt/.xlsx",
                  bg=BTN_BG,fg=C_READ,activebackground=C_READ,activeforeground=BG,
                  relief="flat",font=("Segoe UI",9),padx=8,pady=3,cursor="hand2",bd=0,
                  command=self._cargar_archivo).pack(side="left",padx=4)
        self.lbl_archivo=tk.Label(cfg,text="(sin archivo)",bg=PANEL,fg=TXT_LO,
                                  font=("Courier New",8))
        self.lbl_archivo.pack(side="left",padx=4)

        tk.Frame(self,bg=BORDER,height=1).pack(fill="x",padx=12)

        # Algoritmos
        alg_frame=tk.Frame(self,bg=BG); alg_frame.pack(fill="x",padx=12,pady=8)
        algos=[("🔀 Intercalación",Intercalacion,C_CMP),
               ("⚡ Mezcla Directa",MezclaDirecta,C_AUX),
               ("⚖️ Mezcla Equilibrada",MezclaEquilibrada,C_DONE)]
        self._algo_cls=None; self._algo_btns=[]
        for txt,cls,col in algos:
            b=tk.Button(alg_frame,text=txt,bg=PANEL2,fg=col,
                        activebackground=col,activeforeground=BG,
                        relief="flat",font=("Segoe UI",10,"bold"),
                        padx=14,pady=7,cursor="hand2",bd=0,
                        highlightthickness=2,highlightbackground=BORDER)
            b.pack(side="left",padx=6)
            b.config(command=lambda c=cls,cl=col,bt=b:self._sel_algo(c,cl,bt))
            self._algo_btns.append((b,cls,col))

        tk.Frame(alg_frame,bg=BORDER,width=1).pack(side="left",padx=10,fill="y")
        self.btn_run=tk.Button(alg_frame,text="▶ EJECUTAR",
                               bg=C_DONE,fg=BG,activebackground="#16a34a",activeforeground=BG,
                               relief="flat",font=("Segoe UI",10,"bold"),
                               padx=14,pady=7,cursor="hand2",bd=0,state="disabled",
                               command=self._ejecutar)
        self.btn_run.pack(side="left",padx=4)
        self.btn_pause=tk.Button(alg_frame,text="⏸ Pausar",
                                 bg=BTN_BG,fg=TXT_LO,activebackground=C_WARN,activeforeground=BG,
                                 relief="flat",font=("Segoe UI",10),
                                 padx=10,pady=7,cursor="hand2",bd=0,state="disabled",
                                 command=self._pausar)
        self.btn_pause.pack(side="left",padx=4)
        self.btn_stop=tk.Button(alg_frame,text="■ Detener",
                                bg=BTN_BG,fg=TXT_LO,activebackground=C_ERR,activeforeground=BG,
                                relief="flat",font=("Segoe UI",10),
                                padx=10,pady=7,cursor="hand2",bd=0,state="disabled",
                                command=self._detener)
        self.btn_stop.pack(side="left",padx=4)
        self.lbl_estado=tk.Label(alg_frame,text="● Selecciona un algoritmo",
                                 bg=BG,fg=TXT_LO,font=("Segoe UI",9))
        self.lbl_estado.pack(side="right",padx=10)

        tk.Frame(self,bg=BORDER,height=1).pack(fill="x",padx=12)

        # Barras
        self.barras=BarrasCanvas(self); self.barras.pack(fill="x",padx=12,pady=8)

        # Tarjetas de cintas
        cintas_frame=tk.Frame(self,bg=BG); cintas_frame.pack(fill="x",padx=12,pady=(0,6))
        tk.Label(cintas_frame,text="Hojas en cintas.xlsx:",bg=BG,fg=TXT_MID,
                 font=("Courier New",9,"bold")).pack(anchor="w",pady=(0,4))
        self._tarjetas={}
        grid=tk.Frame(cintas_frame,bg=BG); grid.pack(fill="x")
        cintas_def=[("datos","#64748b"),("cinta_A",CINTA_COLORES["A"]),
                    ("cinta_B",CINTA_COLORES["B"]),("cinta_C",CINTA_COLORES["C"]),
                    ("cinta_D",CINTA_COLORES["D"]),("resultado","#22c55e")]
        for ci,(nombre,color) in enumerate(cintas_def):
            t=TarjetaCinta(grid,nombre,color)
            t.grid(row=0,column=ci,sticky="nsew",padx=3,pady=2)
            grid.columnconfigure(ci,weight=1)
            self._tarjetas[nombre]=t

        # Etiqueta hojas resultado
        res_lbl=tk.Frame(self,bg=BG); res_lbl.pack(fill="x",padx=12)
        tk.Label(res_lbl,
                 text="📋 resultado.xlsx → Hoja 1: Vertical  |  Hoja 2: Horizontal  |  Hoja 3: Revuelto",
                 bg=BG,fg=C_DONE,font=("Courier New",9,"bold")).pack(anchor="w")

        tk.Frame(self,bg=BORDER,height=1).pack(fill="x",padx=12,pady=(4,0))

        # Log
        log_frame=tk.Frame(self,bg=BG); log_frame.pack(fill="both",expand=True,padx=12,pady=6)
        lh=tk.Frame(log_frame,bg=BG); lh.pack(fill="x")
        tk.Label(lh,text="Log de operaciones:",bg=BG,fg=TXT_MID,
                 font=("Courier New",9,"bold")).pack(side="left")
        tk.Button(lh,text="🗑 Limpiar",bg=BTN_BG,fg=TXT_LO,relief="flat",
                  font=("Courier New",8),padx=6,pady=2,cursor="hand2",bd=0,
                  command=self._limpiar_log).pack(side="right")
        self.log_txt=scrolledtext.ScrolledText(log_frame,height=8,
                                               bg="#060a0f",fg=TXT_MID,
                                               font=("Courier New",9),state="disabled",
                                               relief="flat",insertbackground=TXT_HI)
        self.log_txt.pack(fill="both",expand=True,pady=4)
        for tag,color in [("cian",C_CMP),("lila",C_AUX),("verde",C_DONE),
                          ("rojo",C_ERR),("naranja",C_READ),("gris",TXT_LO),
                          ("blanco",TXT_HI),("mid",TXT_MID)]:
            self.log_txt.tag_config(tag,foreground=color)

        pie=tk.Frame(self,bg=BORDER,height=22); pie.pack(fill="x",side="bottom")
        pie.pack_propagate(False)
        tk.Label(pie,text=f"Carpeta: {carpeta_cintas()}",
                 font=("Courier New",8),bg=BORDER,fg=TXT_LO
                 ).pack(side="left",padx=10,pady=2)

    def _vel_changed(self,v):
        self._delay=max(0.02, 1.5-(int(v)-1)/99*1.48)

    def _sel_algo(self,cls,color,btn=None):
        self._algo_cls=cls
        for b,c,col in self._algo_btns:
            b.config(highlightbackground=BORDER,bg=PANEL2)
        if btn: btn.config(highlightbackground=color,bg="#1a2535")
        self.btn_run.config(state="normal")
        self.lbl_estado.config(text=f"● {cls.__name__} seleccionada",fg=color)

    def _cargar_archivo(self):
        path=filedialog.askopenfilename(
            title="Seleccionar archivo",
            filetypes=[("Texto/Excel","*.txt *.xlsx"),("Todos","*.*")])
        if not path: return
        try:
            nums=[]
            if path.endswith(".xlsx"):
                wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
                ws=wb.active
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        try: nums.append(int(cell))
                        except: pass
            else:
                with open(path) as f:
                    for tok in f.read().replace(","," ").replace(";"," ").split():
                        try: nums.append(int(tok))
                        except: pass
            if not nums:
                self._log("⚠ Sin números válidos en el archivo",C_ERR); return
            self._datos_externos=nums
            self.n_var.set(str(len(nums)))
            self.lbl_archivo.config(text=f"✔ {os.path.basename(path)} ({len(nums)})",fg=C_DONE)
            self._log(f"✦ Cargado: {len(nums)} números",C_READ)
        except Exception as e:
            self._log(f"⚠ Error: {e}",C_ERR)

    def _ejecutar(self):
        if self._thread and self._thread.is_alive(): return
        try:
            n=int(self.n_var.get())
            if n<2: self._log("⚠ Mínimo 2",C_ERR); return
        except ValueError:
            self._log("⚠ Número inválido",C_ERR); return

        for t in self._tarjetas.values(): t.limpiar()
        self.barras.dibujar([])
        self._pausado.set()
        self.btn_run.config(state="disabled")
        self.btn_pause.config(state="normal",text="⏸ Pausar")
        self.btn_stop.config(state="normal")

        algo=self._algo_cls(
            log_cb=self._log, barra_cb=self._cb_barra,
            cinta_cb=self._cb_cinta, fin_cb=self._cb_fin,
            pausa_fn=self._pausado.wait)
        self._algo=algo

        def _run():
            try:
                algo.ejecutar(n, self._delay)
            except InterruptedError:
                self._log("\n■ Detenido.",C_ERR)
                self.btn_run.config(state="normal")
                self.btn_pause.config(state="disabled")
                self.btn_stop.config(state="disabled")
            except Exception as e:
                self._log(f"\n⚠ Error: {e}",C_ERR)
                self.btn_run.config(state="normal")
                self.btn_pause.config(state="disabled")
                self.btn_stop.config(state="disabled")

        self._thread=threading.Thread(target=_run,daemon=True)
        self._thread.start()

    def _pausar(self):
        if self._pausado.is_set():
            self._pausado.clear()
            self.btn_pause.config(text="▶ Continuar")
            self.lbl_estado.config(text="● Pausado",fg=C_WARN)
        else:
            self._pausado.set()
            self.btn_pause.config(text="⏸ Pausar")
            self.lbl_estado.config(text="● Ejecutando…",fg=C_CMP)

    def _detener(self):
        if self._algo: self._pausado.set(); self._algo.detener()
        self.btn_run.config(state="normal")
        self.btn_pause.config(state="disabled",text="⏸ Pausar")
        self.btn_stop.config(state="disabled")

    def _log(self,msg,color=None):
        tag={C_CMP:"cian",C_AUX:"lila",C_DONE:"verde",C_ERR:"rojo",
             C_READ:"naranja",TXT_LO:"gris",TXT_HI:"blanco",TXT_MID:"mid",
             C_SWP:"rojo",C_WARN:"naranja"}.get(color,"mid")
        def _do():
            self.log_txt.config(state="normal")
            self.log_txt.insert("end",msg+"\n",tag)
            self.log_txt.see("end")
            self.log_txt.config(state="disabled")
        self.after(0,_do)

    def _cb_barra(self,arr,hi):
        self.after(0,lambda:self.barras.dibujar(arr,hi))

    def _cb_cinta(self,cintas_dict):
        def _do():
            for nombre,datos in cintas_dict.items():
                key=nombre.replace(".txt","").replace(".xlsx","")
                if key in self._tarjetas:
                    self._tarjetas[key].actualizar(datos)
        self.after(0,_do)

    def _cb_fin(self,resultado,path):
        def _do():
            self.lbl_estado.config(text="✔ Completado — resultado.xlsx listo",fg=C_DONE)
            self.btn_run.config(state="normal")
            self.btn_pause.config(state="disabled",text="⏸ Pausar")
            self.btn_stop.config(state="disabled")
            self.barras.dibujar(resultado,{i:C_DONE for i in range(len(resultado))})
            self._log(f"✔ {len(resultado)} elementos ordenados",C_DONE)
            self._log(f"   📊 Hoja 1 'Vertical'   → columna ordenada ascendente",C_AUX)
            self._log(f"   📊 Hoja 2 'Horizontal' → fila ordenada ascendente",C_AUX)
            self._log(f"   📊 Hoja 3 'Revuelto'   → cuadrícula desordenada",C_AUX)
            self._log(f"   Ruta: {path}",TXT_LO)
        self.after(0,_do)

    def _limpiar_log(self):
        self.log_txt.config(state="normal")
        self.log_txt.delete("1.0","end")
        self.log_txt.config(state="disabled")


if __name__=="__main__":
    App().mainloop()