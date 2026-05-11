"""
╔══════════════════════════════════════════════════════════════════╗
║         lib_ordenamiento_externo.py                             ║
║  Librería de Métodos de Ordenamiento Externo                    ║
║                                                                  ║
║  Simula cintas magnéticas usando archivos reales en disco.      ║
║  Soporta DOS modos de E/S intercambiables:                      ║
║    • MODO TXT   → archivos .txt  (sin dependencias extra)       ║
║    • MODO EXCEL → libro cintas.xlsx con hojas por cinta         ║
║                   + resultado.xlsx con 3 presentaciones         ║
║                                                                  ║
║  Métodos incluidos:                                              ║
║    1. Intercalación      (Straight Merge)                       ║
║    2. Mezcla Directa     (Top-Down Merge Sort)                  ║
║    3. Mezcla Equilibrada (Balanced 4-tape Merge)                ║
║                                                                  ║
║  Cada algoritmo hereda de AlgoritmoBase y recibe callbacks      ║
║  para log, barras visuales, cintas y finalización.              ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import random
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

METODOS_EXTERNOS = [
    "Intercalación",
    "Mezcla Directa",
    "Mezcla Equilibrada",
]

MODOS_IO = ["TXT", "EXCEL"]

C_CMP  = "#00d4ff"
C_SWP  = "#ff6b6b"
C_AUX  = "#a78bfa"
C_DONE = "#22c55e"
C_READ = "#f59e0b"
TXT_LO = "#64748b"
TXT_MID= "#94a3b8"
TXT_HI = "#e2e8f0"
C_WARN = "#f59e0b"
C_ERR  = "#ff6b6b"

CINTA_COLORES = {
    "A": "#00d4ff",
    "B": "#a78bfa",
    "C": "#22c55e",
    "D": "#f59e0b",
}

_XLSX_FILLS = {
    "cinta_A": "D6F0FB", "cinta_B": "EDE9FF",
    "cinta_C": "D4F4E2", "cinta_D": "FEF3C7",
    "datos":   "F1F5F9", "entrada": "F1F5F9",
    "temp1":   "FFF1F0", "temp2":   "FFF8EC",
}

def _mk_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color) \
           if OPENPYXL_OK else None

def carpeta_cintas(base_dir=None):
    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    ruta = os.path.join(base_dir, "cintas")
    os.makedirs(ruta, exist_ok=True)
    return ruta

def _ruta(nombre, carpeta):
    return os.path.join(carpeta, nombre)

def escribir_cinta_txt(nombre, numeros, carpeta):
    with open(_ruta(nombre + ".txt", carpeta), "w") as f:
        for n in numeros:
            f.write(f"{n}\n")

def leer_cinta_txt(nombre, carpeta):
    path = _ruta(nombre + ".txt", carpeta)
    if not os.path.exists(path):
        return []
    with open(path, "r") as f:
        lineas = f.read().strip().split("\n")
    return [int(x) for x in lineas if x.strip()]

def tamanio_cinta_txt(nombre, carpeta):
    path = _ruta(nombre + ".txt", carpeta)
    return os.path.getsize(path) if os.path.exists(path) else 0

def _get_libro(carpeta):
    path = _ruta("cintas.xlsx", carpeta)
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception:
            pass
    return openpyxl.Workbook()

def escribir_cinta_excel(nombre, numeros, carpeta):
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado.")
    path = _ruta("cintas.xlsx", carpeta)
    wb   = _get_libro(carpeta)
    hoja = nombre
    if hoja in wb.sheetnames:
        del wb[hoja]
    ws = wb.create_sheet(hoja)
    fill_hex = _XLSX_FILLS.get(hoja, "F8FAFC")
    header_fill = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=9)
    thin        = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    for col, val in [("A", "#"), ("B", "Valor")]:
        c = ws[f"{col}1"]
        c.value = val; c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center"); c.border = cell_border
    row_fill = _mk_fill(fill_hex)
    for i, val in enumerate(numeros, start=1):
        ws[f"A{i+1}"] = i; ws[f"B{i+1}"] = val
        for col2 in ("A", "B"):
            c2 = ws[f"{col2}{i+1}"]
            c2.fill = row_fill; c2.font = cell_font
            c2.border = cell_border; c2.alignment = Alignment(horizontal="center")
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    wb.save(path)

def leer_cinta_excel(nombre, carpeta):
    if not OPENPYXL_OK:
        return []
    path = _ruta("cintas.xlsx", carpeta)
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if nombre not in wb.sheetnames:
        return []
    out = []
    for row in wb[nombre].iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            try:
                out.append(int(row[1]))
            except (ValueError, TypeError):
                pass
    return out

def tamanio_cinta_excel(carpeta):
    path = _ruta("cintas.xlsx", carpeta)
    return os.path.getsize(path) if os.path.exists(path) else 0

def escribir_resultado_excel(numeros, carpeta):
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado.")
    path = _ruta("resultado.xlsx", carpeta)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)
    n = len(numeros)
    header_fill = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=9)
    thin        = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws1 = wb.create_sheet("Vertical")
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16
    for col, val in [("A", "#"), ("B", "Valor ordenado")]:
        c = ws1[f"{col}1"]
        c.value = val; c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center"); c.border = cell_border
    greens = ["D4F4E2","C6EFD3","B7E9C4","A8E3B5","99DDA6"]
    for i, val in enumerate(numeros, start=1):
        fill_hex = greens[min(int(i / max(n, 1) * len(greens)), len(greens)-1)]
        ws1[f"A{i+1}"] = i; ws1[f"B{i+1}"] = val
        for col2 in ("A", "B"):
            c2 = ws1[f"{col2}{i+1}"]
            c2.fill = _mk_fill(fill_hex); c2.font = cell_font
            c2.border = cell_border; c2.alignment = Alignment(horizontal="center")
    ws1.freeze_panes = "A2"
    ws2 = wb.create_sheet("Horizontal")
    blues = ["D6EAF8","C0D8F0","AAC6E8","94B4E0","7EA2D8"]
    for col_idx, val in enumerate(numeros, start=1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 8
        h = ws2.cell(row=1, column=col_idx, value=col_idx)
        h.fill = header_fill; h.font = header_font
        h.alignment = Alignment(horizontal="center"); h.border = cell_border
        fill_hex = blues[min(int(col_idx / max(n, 1) * len(blues)), len(blues)-1)]
        v = ws2.cell(row=2, column=col_idx, value=val)
        v.fill = _mk_fill(fill_hex); v.font = cell_font
        v.border = cell_border; v.alignment = Alignment(horizontal="center")
    ws3 = wb.create_sheet("Revuelto")
    cols = max(1, min(20, int(n**0.5) + 1))
    rows = (n + cols - 1) // cols
    revuelto = numeros[:]
    random.shuffle(revuelto)
    ws3.merge_cells("A1:" + get_column_letter(cols) + "1")
    t = ws3["A1"]
    t.value = "Datos Revueltos (desordenados)"
    t.fill  = _mk_fill("7C3AED")
    t.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    t.alignment = Alignment(horizontal="center"); t.border = cell_border
    paleta = ["FFF3CD","D4EDDA","D1ECF1","F8D7DA","E2D9F3",
              "FDE2E4","C3E6CB","BEE5EB","FFEEBA","D6E4FF"]
    idx = 0
    for r in range(rows):
        for c in range(cols):
            if idx >= len(revuelto): break
            col_letter = get_column_letter(c + 1)
            ws3.column_dimensions[col_letter].width = 9
            cell = ws3.cell(row=r+2, column=c+1, value=revuelto[idx])
            cell.fill = _mk_fill(paleta[idx % len(paleta)])
            cell.font = cell_font; cell.border = cell_border
            cell.alignment = Alignment(horizontal="center")
            idx += 1
        ws3.row_dimensions[r+2].height = 16
    wb.save(path)
    return path


class AlgoritmoBase:
    def __init__(self, log_cb, barra_cb, cinta_cb, fin_cb, pausa_fn,
                 modo_io="TXT", carpeta=None):
        self.log    = log_cb
        self.barra  = barra_cb
        self.cinta  = cinta_cb
        self.fin    = fin_cb
        self.pausa  = pausa_fn
        self.modo   = modo_io.upper()
        self.folder = carpeta or carpeta_cintas()
        self._stop  = False

    def detener(self):
        self._stop = True

    def _chk(self):
        if self._stop:
            raise InterruptedError
        self.pausa()

    def _escribir(self, nombre, datos):
        if self.modo == "EXCEL":
            escribir_cinta_excel(nombre, datos, self.folder)
        else:
            escribir_cinta_txt(nombre, datos, self.folder)

    def _leer(self, nombre):
        if self.modo == "EXCEL":
            return leer_cinta_excel(nombre, self.folder)
        else:
            return leer_cinta_txt(nombre, self.folder)

    def _tamanio(self, nombre):
        if self.modo == "EXCEL":
            return tamanio_cinta_excel(self.folder)
        else:
            return tamanio_cinta_txt(nombre, self.folder)

    def _nombre_archivo(self, nombre):
        if self.modo == "EXCEL":
            return f"hoja '{nombre}'"
        return f"{nombre}.txt"

    def _mostrar_cintas(self, *nombres):
        d = {}
        for n in nombres:
            d[n] = self._leer(n)
        self.cinta(d)

    def _guardar_resultado(self, data):
        if self.modo == "EXCEL":
            path = escribir_resultado_excel(data, self.folder)
            return path
        else:
            self._escribir("resultado", data)
            return os.path.join(self.folder, "resultado.txt")


class Intercalacion(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        modo = "Excel" if self.modo == "EXCEL" else "TXT"
        self.log("═" * 50, TXT_LO)
        self.log(f"INTERCALACIÓN [{modo}]", C_CMP)
        self.log("═" * 50, TXT_LO)
        datos = random.choices(range(1, n_elem * 10 + 1), k=n_elem)
        self._escribir("datos", datos)
        self.log(f"✦ {self._nombre_archivo('datos')} generada ({n_elem} elementos)", C_READ)
        self._mostrar_cintas("datos")
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()
        data, size, n, paso = datos[:], 1, len(datos), 0
        while size < n:
            self._chk()
            paso += 1
            self.log(f"\n── Pasada {paso} (bloque={size}) ──", TXT_MID)
            cA, cB, i, toggle = [], [], 0, True
            while i < n:
                bloque = data[i:i + size]
                (cA if toggle else cB).extend(bloque)
                toggle = not toggle
                i += size
            self._escribir("cinta_A", cA)
            self._escribir("cinta_B", cB)
            self.log(f"  → cinta_A ({len(cA)}) | cinta_B ({len(cB)})", C_AUX)
            self._mostrar_cintas("cinta_A", "cinta_B")
            time.sleep(delay)
            self._chk()
            resultado, left = [], 0
            while left < n:
                mid   = min(left + size,     n)
                right = min(left + size * 2, n)
                i2, j = left, mid
                while i2 < mid and j < right:
                    self._chk()
                    self.barra(data, {i2: C_CMP, j: C_CMP})
                    time.sleep(delay * 0.3)
                    if data[i2] <= data[j]:
                        resultado.append(data[i2]); i2 += 1
                    else:
                        resultado.append(data[j]);  j  += 1
                while i2 < mid:
                    resultado.append(data[i2]); i2 += 1
                while j < right:
                    resultado.append(data[j]);  j  += 1
                left += size * 2
            data = resultado
            self._escribir("datos", data)
            self.log(f"  ✔ fusión → {self._nombre_archivo('datos')} actualizada", C_DONE)
            self.barra(data, {})
            self._mostrar_cintas("datos")
            time.sleep(delay)
            size *= 2
        path = self._guardar_resultado(data)
        self.log(f"\n✔ ORDENADO → {path}", C_DONE)
        self._mostrar_cintas("datos")
        self.fin(data, path)


class MezclaDirecta(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        modo = "Excel" if self.modo == "EXCEL" else "TXT"
        self.log("═" * 50, TXT_LO)
        self.log(f"MEZCLA DIRECTA [{modo}]", C_CMP)
        self.log("═" * 50, TXT_LO)
        datos = random.choices(range(1, n_elem * 10 + 1), k=n_elem)
        entrada_nombre = "datos" if self.modo == "EXCEL" else "entrada"
        self._escribir(entrada_nombre, datos)
        self.log(f"✦ {self._nombre_archivo(entrada_nombre)} generada ({n_elem} elementos)", C_READ)
        self._mostrar_cintas(entrada_nombre)
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()
        data, steps = datos[:], []

        def merge(lo, mid, hi):
            left  = data[lo:mid + 1]
            right = data[mid + 1:hi + 1]
            i = j = 0; k = lo
            while i < len(left) and j < len(right):
                if left[i] <= right[j]:
                    data[k] = left[i]; i += 1
                else:
                    data[k] = right[j]; j += 1
                steps.append(("merge", data[:], lo, mid, hi, k)); k += 1
            while i < len(left):
                data[k] = left[i]; i += 1; k += 1
                steps.append(("merge", data[:], lo, mid, hi, k - 1))
            while j < len(right):
                data[k] = right[j]; j += 1; k += 1
                steps.append(("merge", data[:], lo, mid, hi, k - 1))

        def sort(lo, hi):
            if lo >= hi: return
            mid = (lo + hi) // 2
            steps.append(("div", data[:], lo, mid, hi, 0))
            sort(lo, mid); sort(mid + 1, hi); merge(lo, mid, hi)
            steps.append(("merged", data[:], lo, mid, hi, 0))

        sort(0, len(data) - 1)
        for ev, snap, lo, mid, hi, extra in steps:
            self._chk()
            if ev == "div":
                self.barra(snap, {k: C_AUX for k in range(lo, hi + 1)})
                self.log(f"  División [{lo}..{mid}] y [{mid+1}..{hi}]", C_AUX)
                self._escribir("temp1", snap[lo:mid + 1])
                self._escribir("temp2", snap[mid + 1:hi + 1])
                self._mostrar_cintas("temp1", "temp2")
                time.sleep(delay * 0.5)
            elif ev == "merge":
                k_pos = extra
                self.barra(snap, {k_pos: C_SWP, lo: C_CMP, hi: C_CMP})
                self._escribir(entrada_nombre, snap)
                time.sleep(delay * 0.2)
            elif ev == "merged":
                self.barra(snap, {k: C_DONE for k in range(lo, hi + 1)})
                self.log(f"  ✔ Fusión [{lo}..{hi}]", C_DONE)
                self._mostrar_cintas(entrada_nombre)
                time.sleep(delay * 0.4)
        path = self._guardar_resultado(data)
        self.log(f"\n✔ ORDENADO → {path}", C_DONE)
        self.fin(data, path)


class MezclaEquilibrada(AlgoritmoBase):
    def ejecutar(self, n_elem, delay):
        modo = "Excel" if self.modo == "EXCEL" else "TXT"
        self.log("═" * 50, TXT_LO)
        self.log(f"MEZCLA EQUILIBRADA — 4 cintas [{modo}]", C_CMP)
        self.log("═" * 50, TXT_LO)
        datos = random.choices(range(1, n_elem * 10 + 1), k=n_elem)
        self._escribir("datos", datos)
        self.log(f"✦ {self._nombre_archivo('datos')} generada ({n_elem} elementos)", C_READ)
        cA = datos[0::2]; cB = datos[1::2]
        self._escribir("cinta_A", cA)
        self._escribir("cinta_B", cB)
        self.log(f"✦ Distribución inicial:", C_AUX)
        self.log(f"   cinta_A → {len(cA)} elementos", CINTA_COLORES["A"])
        self.log(f"   cinta_B → {len(cB)} elementos", CINTA_COLORES["B"])
        self._mostrar_cintas("cinta_A", "cinta_B")
        self.barra(datos, {})
        time.sleep(delay)
        self._chk()
        data = datos[:]
        n, size, pasada = len(data), 1, 0
        pares = [
            ("cinta_A", "cinta_B", "cinta_C", "cinta_D"),
            ("cinta_C", "cinta_D", "cinta_A", "cinta_B"),
        ]
        while size < n:
            self._chk()
            ent1, ent2, sal1, sal2 = pares[pasada % 2]
            pasada += 1
            self.log(f"\n── Pasada {pasada} (bloque={size})", TXT_MID)
            self.log(f"   Lee : {ent1} + {ent2}", C_READ)
            self.log(f"   Escribe: {sal1} + {sal2}", C_SWP)
            inp1 = self._leer(ent1)
            inp2 = self._leer(ent2)
            data_ent, i1, i2, toggle = [], 0, 0, True
            while i1 < len(inp1) or i2 < len(inp2):
                if toggle and i1 < len(inp1):
                    data_ent.extend(inp1[i1:i1 + size]); i1 += size
                elif not toggle and i2 < len(inp2):
                    data_ent.extend(inp2[i2:i2 + size]); i2 += size
                else:
                    if i1 < len(inp1):
                        data_ent.extend(inp1[i1:i1 + size]); i1 += size
                    elif i2 < len(inp2):
                        data_ent.extend(inp2[i2:i2 + size]); i2 += size
                toggle = not toggle
            out1, out2, left, t2 = [], [], 0, True
            while left < len(data_ent):
                mid2  = min(left + size,     len(data_ent))
                right = min(left + size * 2, len(data_ent))
                ia, jb, merged = left, mid2, []
                while ia < mid2 and jb < right:
                    self._chk()
                    self.barra(data_ent, {ia: C_CMP, jb: C_CMP})
                    time.sleep(delay * 0.15)
                    if data_ent[ia] <= data_ent[jb]:
                        merged.append(data_ent[ia]); ia += 1
                    else:
                        merged.append(data_ent[jb]); jb += 1
                while ia < mid2:
                    merged.append(data_ent[ia]); ia += 1
                while jb < right:
                    merged.append(data_ent[jb]); jb += 1
                (out1 if t2 else out2).extend(merged)
                t2 = not t2; left += size * 2
            resultado, i1, i2, tog = [], 0, 0, True
            while i1 < len(out1) or i2 < len(out2):
                bs = size * 2
                if tog and i1 < len(out1):
                    resultado.extend(out1[i1:i1 + bs]); i1 += bs
                elif not tog and i2 < len(out2):
                    resultado.extend(out2[i2:i2 + bs]); i2 += bs
                else:
                    if i1 < len(out1):   resultado.extend(out1[i1:i1 + bs]); i1 += bs
                    elif i2 < len(out2): resultado.extend(out2[i2:i2 + bs]); i2 += bs
                tog = not tog
            data = resultado or out1 + out2
            self._escribir(sal1, out1)
            self._escribir(sal2, out2)
            self.log(f"   {sal1}: {len(out1)} | {sal2}: {len(out2)}", C_DONE)
            self._mostrar_cintas(sal1, sal2)
            self.barra(data, {})
            time.sleep(delay)
            size *= 2
        ult_sal1 = pares[(pasada - 1) % 2][2]
        ult_sal2 = pares[(pasada - 1) % 2][3]
        final = sorted(self._leer(ult_sal1) + self._leer(ult_sal2))
        path = self._guardar_resultado(final)
        self.log(f"\n✔ ORDENADO → {path}", C_DONE)
        self._mostrar_cintas(ult_sal1, ult_sal2)
        self.barra(final, {i: C_DONE for i in range(len(final))})
        self.fin(final, path)


CLASES_EXTERNAS = {
    "Intercalación":      Intercalacion,
    "Mezcla Directa":     MezclaDirecta,
    "Mezcla Equilibrada": MezclaEquilibrada,
}


def ordenar_externo(nombre_metodo: str, n_elem: int,
                    modo_io: str = "TXT", delay: float = 0.0,
                    carpeta: str = None) -> list:
    if nombre_metodo not in CLASES_EXTERNAS:
        raise ValueError(f"Método desconocido: {nombre_metodo}.")
    resultado_final = []
    def _fin(res, path):
        resultado_final.extend(res)
    algo = CLASES_EXTERNAS[nombre_metodo](
        log_cb=lambda msg, color=None: None,
        barra_cb=lambda arr, hi: None,
        cinta_cb=lambda d: None,
        fin_cb=_fin,
        pausa_fn=lambda: None,
        modo_io=modo_io,
        carpeta=carpeta,
    )
    algo.ejecutar(n_elem, delay)
    return resultado_final


if __name__ == "__main__":
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        for nombre in METODOS_EXTERNOS:
            res = ordenar_externo(nombre, 12, modo_io="TXT",
                                  delay=0.0, carpeta=tmp)
            print(f"{nombre:<22}: {res}")